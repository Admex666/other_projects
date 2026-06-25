import streamlit as st
import pandas as pd
import openpyxl
import os
import json
import re
from dotenv import load_dotenv
from google.oauth2 import service_account
from googleapiclient.discovery import build

# Page config
st.set_page_config(
    page_title="VitaSteps Logisztikai Dashboard",
    page_icon="🏔️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom Styling
st.markdown("""
<style>
    .reportview-container {
        background: #111111;
    }
    h1, h2, h3 {
        font-family: 'Outfit', sans-serif;
    }
    .stMetric {
        background: rgba(255, 255, 255, 0.03);
        border: 1px solid rgba(255, 255, 255, 0.1);
        padding: 1.5rem 1rem;
        border-radius: 8px;
        text-align: center;
    }
    .matrix-card {
        background: rgba(196, 255, 0, 0.02);
        border: 1px solid rgba(196, 255, 0, 0.15);
        border-radius: 8px;
        padding: 1.5rem;
        text-align: center;
    }
    .matrix-header {
        font-weight: bold;
        font-size: 1.1rem;
        color: #c4ff00;
        margin-bottom: 0.8rem;
    }
    .matrix-val {
        font-size: 2.2rem;
        font-weight: 800;
        color: #ffffff;
    }
    .matrix-sub {
        font-size: 0.85rem;
        color: #888888;
        margin-top: 0.5rem;
    }
    .badge {
        padding: 0.2rem 0.6rem;
        border-radius: 4px;
        font-size: 0.8rem;
        font-weight: bold;
    }
    .badge-success {
        background: rgba(40, 167, 69, 0.15);
        color: #28a745;
        border: 1px solid rgba(40, 167, 69, 0.3);
    }
    .badge-warning {
        background: rgba(255, 193, 7, 0.15);
        color: #ffc107;
        border: 1px solid rgba(255, 193, 7, 0.3);
    }
    .badge-danger {
        background: rgba(220, 53, 69, 0.15);
        color: #dc3545;
        border: 1px solid rgba(220, 53, 69, 0.3);
    }
</style>
""", unsafe_allow_html=True)

# Load configuration
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.join(SCRIPT_DIR, '.env')
load_dotenv(ENV_PATH)

SHEET_ID = os.getenv("GOOGLE_SHEET_ID")
SERVICE_ACCOUNT_JSON = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
XLSX_TEMPLATE_PATH = r"C:\Users\Adam\Downloads\Tomeges-import-pelda-hu-HU.xlsx"
OUTPUT_XLSX_PATH = r"C:\Users\Adam\Downloads\Foxpost_import_ready.xlsx"

# Make sure we search for template in Downloads as a backup or default
if not os.path.exists(XLSX_TEMPLATE_PATH):
    # Try local template or absolute downloads path if it was moved
    user_home = os.path.expanduser("~")
    backup_path = os.path.join(user_home, "Downloads", "Tomeges-import-pelda-hu-HU.xlsx")
    if os.path.exists(backup_path):
        XLSX_TEMPLATE_PATH = backup_path
    else:
        # Fallback to a relative location in project if any
        proj_backup = os.path.join(SCRIPT_DIR, "Tomeges-import-pelda-hu-HU.xlsx")
        if os.path.exists(proj_backup):
            XLSX_TEMPLATE_PATH = proj_backup

SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/spreadsheets.readonly"]

# Caching for API calls and Excel loading
@st.cache_data
def fetch_sheet_rows(sheet_name):
    try:
        creds_dict = json.loads(SERVICE_ACCOUNT_JSON)
        creds = service_account.Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
        service = build("sheets", "v4", credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=SHEET_ID,
            range=f"{sheet_name}!A1:AJ500"
        ).execute()
        return result.get("values", [])
    except Exception as e:
        st.error(f"Hiba a Google Sheets lekérésekor ({sheet_name}): {e}")
        return []

@st.cache_data
def load_foxpost_lockers(template_path):
    try:
        wb = openpyxl.load_workbook(template_path, read_only=True)
        sheet = wb['Automata adatok']
        rows = list(sheet.iter_rows(values_only=True))
        
        lockers = []
        for r in rows[1:]:
            if len(r) >= 2 and r[0] and r[1]:
                name = str(r[0]).strip()
                code = str(r[1]).strip()
                lockers.append((name, code))
        return lockers
    except Exception as e:
        st.error(f"Hiba a Foxpost template betöltésekor ({template_path}): {e}")
        return []

def clean_phone(phone_str):
    if not phone_str:
        return ""
    digits = "".join(c for c in str(phone_str) if c.isdigit())
    if not digits:
        return ""
    if digits.startswith("06") and len(digits) == 11:
        return "36" + digits[2:]
    if digits.startswith("36") and len(digits) == 11:
        return digits
    if len(digits) == 9 and digits[0] in ("2", "3", "7", "5"):
        return "36" + digits
    return digits

def clean_locker_input(text):
    if not text:
        return ""
    main_part = text.split('(')[0].strip()
    return main_part

def resolve_locker_code(point_text, lockers_list):
    if not point_text:
        return None
    clean_pt = clean_locker_input(point_text)
    
    input_tokens = set(re.findall(r'[a-z0-9űáéúőóüöí]+', clean_pt.lower()))
    stop_words = {"foxpost", "abox", "zbox", "automatak", "automata", "ker", "bp", "budapest", "utca", "ut", "ter", "u", "szam", "de", "es", "a", "az", "kerulet"}
    input_tokens = {t for t in input_tokens if t not in stop_words and len(t) > 1}
    
    best_match_code = None
    best_match_score = 0
    
    for official_name, code in lockers_list:
        off_tokens = set(re.findall(r'[a-z0-9űáéúőóüöí]+', official_name.lower()))
        off_tokens = {t for t in off_tokens if t not in stop_words and len(t) > 1}
        
        score = len(input_tokens.intersection(off_tokens))
        if score > best_match_score:
            best_match_score = score
            best_match_code = code
            
    if best_match_score >= 1:
        return best_match_code
    return None

def share_family_name(name1, name2):
    p1 = name1.strip().split()
    p2 = name2.strip().split()
    if not p1 or not p2:
        return False
    f1 = p1[0].lower()
    f2 = p2[0].lower()
    if f1 == f2 or f1 in f2 or f2 in f1:
        return True
    return False

def main():
    st.title("🏔️ VitaSteps Logisztikai Dashboard & Elemző")
    st.markdown("---")
    
    # Check setup files
    if not os.path.exists(XLSX_TEMPLATE_PATH):
        st.error(f"A megadott Foxpost template Excel nem található ezen az útvonalon: `{XLSX_TEMPLATE_PATH}`. Kérjük, másold oda, vagy győződj meg a letöltési mappádról!")
        return

    # Data loading
    with st.spinner("Adatok lekérése a Google Sheets-ből és Excel-ből..."):
        nevezesek_rows = fetch_sheet_rows("Nevezések")
        tally_rows = fetch_sheet_rows("tally_szallitas")
        lockers_list = load_foxpost_lockers(XLSX_TEMPLATE_PATH)
        
    if not nevezesek_rows or not tally_rows:
        st.error("Nem sikerült beolvasni a Google Sheets táblákat.")
        return

    # 1. Parse Tally Shipping Submissions
    tally_map = {}
    tally_headers = tally_rows[0]
    for row in tally_rows[1:]:
        row += [""] * (len(tally_headers) - len(row))
        tally_email1 = row[4].strip().lower()
        tally_email2 = row[7].strip().lower()
        
        tally_data = {
            "name": row[3] or row[6],
            "email": row[4] or row[7],
            "phone": clean_phone(row[8]),
            "point": row[5] or row[9],
            "comment": row[12]
        }
        if tally_email1:
            tally_map[tally_email1] = tally_data
        if tally_email2:
            tally_map[tally_email2] = tally_data

    # 2. Parse Nevezések
    nev_headers = nevezesek_rows[0]
    
    def find_col(name, default):
        for idx, h in enumerate(nev_headers):
            if h.strip().lower() == name.lower().strip():
                return idx
        return default

    col_email = find_col("email", 3)
    col_nev = find_col("név", 4)
    col_teljesitve = find_col("teljesítve dátum", 12)
    col_szallitas_tip = find_col("szállítás típus", 19)
    col_szallitasi_cim = find_col("szállítási cím", 20)
    col_szallitasi_tel = find_col("szállítási telefonszám", 21)
    col_ermek_szama = find_col("érmek száma", 22)
    col_kikuldve = find_col("érem kiküldve?", 24)
    col_comment = find_col("megjegyzés", 27)
    col_egyutt_kuldve = find_col("együtt küldve", 26)
    
    raw_runners = []
    
    for idx, row in enumerate(nevezesek_rows[1:], start=2):
        row += [""] * (len(nev_headers) - len(row))
        email = row[col_email].strip().lower()
        name = row[col_nev].strip()
        teljesitve = row[col_teljesitve].strip()
        kikuldve = row[col_kikuldve].strip().lower()
        szallitasi_cim = row[col_szallitasi_cim].strip()
        szallitasi_tel = clean_phone(row[col_szallitasi_tel].strip())
        comment = row[col_comment].strip()
        egyutt_kuldve = row[col_egyutt_kuldve].strip() if col_egyutt_kuldve < len(row) else ""
        
        if not email or not name:
            continue
            
        # Try to gather shipping info
        phone = szallitasi_tel
        point = szallitasi_cim
        tally_comment = ""
        
        if email in tally_map:
            t_data = tally_map[email]
            if not phone:
                phone = t_data["phone"]
            if not point:
                point = t_data["point"]
            tally_comment = t_data["comment"]
            
        resolved_code = resolve_locker_code(point, lockers_list)
        
        ermek_szama_val = row[col_ermek_szama].strip() if col_ermek_szama < len(row) else ""
        try:
            ermek_szama = int(ermek_szama_val) if (ermek_szama_val and str(ermek_szama_val).isdigit()) else 1
        except ValueError:
            ermek_szama = 1
        
        raw_runners.append({
            "row_index": idx,
            "name": name,
            "email": email,
            "phone": phone,
            "point_description": point,
            "locker_code": resolved_code,
            "completed": bool(teljesitve),
            "completion_date": teljesitve,
            "shipped": kikuldve in ("igen", "yes"),
            "comment": comment,
            "tally_comment": tally_comment,
            "egyutt_kuldve": egyutt_kuldve,
            "ermek_szama": ermek_szama
        })

    # Resolution pass for combined shipments
    all_runners = []
    for r in raw_runners:
        phone = r["phone"]
        point = r["point_description"]
        resolved_code = r["locker_code"]
        egyutt_val = r["egyutt_kuldve"].strip()
        primary_email = ""
        is_sub_order = False
        
        # Try to find a primary runner to inherit details from
        primary = None
        if egyutt_val:
            primary = next((x for x in raw_runners if x["email"].lower() == egyutt_val.lower()), None)
            if not primary:
                primary = next((x for x in raw_runners if x["name"].lower() == egyutt_val.lower()), None)
            if primary:
                primary_email = primary["email"]
                is_sub_order = True
        else:
            # If not explicitly linked, find another runner with the same email who has details filled
            primary = next((x for x in raw_runners if x["email"].lower() == r["email"].lower() and x["row_index"] != r["row_index"] and (x["phone"] or x["point_description"])), None)
            if primary:
                primary_email = primary["email"]
                is_sub_order = True
                
        if primary:
            if not phone:
                phone = primary["phone"]
            if not point or point.lower() in ("", "#n/a", "#name?", "#value!"):
                point = primary["point_description"]
            if not resolved_code:
                resolved_code = primary["locker_code"]
        
        has_phone = bool(phone)
        has_point = bool(point and point.lower() not in ("", "#n/a", "#name?", "#value!"))
        has_locker_code = bool(resolved_code)
        
        all_data_ok = has_phone and has_point and has_locker_code
        
        missing = []
        if not has_phone:
            missing.append("telefonszám")
        if not has_point:
            missing.append("csomagpont")
        elif not has_locker_code:
            missing.append("hibás csomagpont kód")
            
        all_runners.append({
            "row_index": r["row_index"],
            "name": r["name"],
            "email": r["email"],
            "phone": phone,
            "point_description": point,
            "locker_code": resolved_code,
            "completed": r["completed"],
            "completion_date": r["completion_date"],
            "shipped": r["shipped"],
            "all_data_ok": all_data_ok,
            "missing_fields": missing,
            "comment": r["comment"],
            "tally_comment": r["tally_comment"],
            "primary_buyer": primary_email,
            "is_sub_order": is_sub_order,
            "egyutt_kuldve": r["egyutt_kuldve"],
            "ermek_szama": r["ermek_szama"]
        })

    # Segmenting runners
    # Group 1: Completed + All data OK
    group1 = [r for r in all_runners if r["completed"] and r["all_data_ok"]]
    # Group 2: Completed + Incomplete data
    group2 = [r for r in all_runners if r["completed"] and not r["all_data_ok"]]
    # Group 3: Not completed + Has data (at least phone or point)
    group3 = [r for r in all_runners if not r["completed"] and (r["phone"] or r["point_description"])]
    # Group 4: Not completed + No data
    group4 = [r for r in all_runners if not r["completed"] and not (r["phone"] or r["point_description"])]
    
    # Build group shipments for display
    all_groups = {}
    for r in all_runners:
        key = r["primary_buyer"] if (r["is_sub_order"] and r["primary_buyer"]) else r["email"]
        if key not in all_groups:
            primary_runner = next((x for x in raw_runners if x["email"].lower() == key.lower()), None)
            if primary_runner:
                recip_name = primary_runner["name"]
                recip_email = primary_runner["email"]
                recip_phone = primary_runner["phone"]
                point_desc = primary_runner["point_description"]
                locker_code = primary_runner["locker_code"]
            else:
                recip_name = r["name"]
                recip_email = r["email"]
                recip_phone = r["phone"]
                point_desc = r["point_description"]
                locker_code = r["locker_code"]
                
            all_groups[key] = {
                "recipient_name": recip_name,
                "recipient_email": recip_email,
                "recipient_phone": recip_phone,
                "point_description": point_desc,
                "locker_code": locker_code,
                "members": []
            }
        all_groups[key]["members"].append(r)

    package_rows = []
    code_to_name = {code: name for name, code in lockers_list}
    
    for key, g in all_groups.items():
        members = g["members"]
        planned_medals = sum(m["ermek_szama"] for m in members)
        
        # Count only members who are completed (shipped or unshipped)
        completed_members = [m for m in members if m["completed"]]
        completed_count = len(completed_members)
        
        # Members who are completed and not shipped yet
        completed_unshipped = [m for m in completed_members if not m["shipped"]]
        completed_unshipped_count = len(completed_unshipped)
        
        # Show package if there's at least one completed runner in the group
        if completed_count > 0:
            member_desc_list = []
            for m in members:
                status_icon = "🟢" if m["completed"] else "🔴"
                shipped_status = " (Feladva)" if m["shipped"] else ""
                member_desc_list.append(f"{m['name']}: {status_icon} {'Teljesített' if m['completed'] else 'Még nem'}{shipped_status} [{m['ermek_szama']} érem]")
                
            members_status_str = " | ".join(member_desc_list)
            
            # Group recipient name for shipment (dedup and only show display name logic)
            recip_name = g["recipient_name"]
            unique_runners = []
            seen_runners = set()
            for runner in members:
                if runner["completed"] and not runner["shipped"]:
                    r_key = (runner["name"].lower().strip(), runner["email"].lower().strip())
                    if r_key not in seen_runners:
                        seen_runners.add(r_key)
                        unique_runners.append(runner)
            
            display_runners = [x["name"] for x in unique_runners if x["email"].lower().strip() != g["recipient_email"].lower().strip()]
            if display_runners:
                display_recip_name = f"{recip_name} (+ {', '.join(display_runners)})"
            else:
                display_recip_name = recip_name
                
            shipping_ok = bool(g["recipient_phone"] and g["recipient_email"] and g["locker_code"])
            status_badge = "🟢 Kész" if shipping_ok else "❌ Hiányos"
            
            off_name = "Ismeretlen kód"
            if g["locker_code"]:
                off_name = code_to_name.get(g["locker_code"], "Ismeretlen kód")
                
            package_rows.append({
                "Címzett": display_recip_name,
                "Recipient Email": g["recipient_email"],
                "Telefonszám": g["recipient_phone"] or "Hiányzik",
                "Csomagpont": g["point_description"] or "Hiányzik",
                "Feloldott Kód": g["locker_code"] or "Hiányzik",
                "Hivatalos Csomagpont Név": off_name,
                "Küldendő érmek (Most)": completed_unshipped_count,
                "Tervezett érmek (Összes)": planned_medals,
                "Csoporttagok státusza": members_status_str,
                "Státusz": status_badge
            })
    
    # 2x2 Matrix Section
    st.header("📊 2x2 Ügyfél Szegmentációs Mátrix")
    
    # Styling columns
    m_col1, m_col2 = st.columns(2)
    
    with m_col1:
        st.markdown(f"""
        <div class="matrix-card">
            <div class="matrix-header">🟢 Teljesített + Minden adat megvan</div>
            <div class="matrix-val">{len(group1)} fő</div>
            <div class="matrix-sub">Ők teljesítették a távot és minden adatuk (tel., cím, kód) megvan a postázáshoz.</div>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<div style='margin-top: 1.5rem;'></div>", unsafe_allow_html=True)
        
        st.markdown(f"""
        <div class="matrix-card" style="border-color: rgba(255,255,255,0.08); background: rgba(255,255,255,0.01);">
            <div class="matrix-header" style="color: #cccccc;">⚪ Nem teljesített + Van adata</div>
            <div class="matrix-val" style="color: #dddddd;">{len(group3)} fő</div>
            <div class="matrix-sub">Ők még nem fejezték be a futást, de már adtak meg szállítási adatokat.</div>
        </div>
        """, unsafe_allow_html=True)
        
    with m_col2:
        st.markdown(f"""
        <div class="matrix-card" style="background: rgba(255, 193, 7, 0.02); border-color: rgba(255, 193, 7, 0.25);">
            <div class="matrix-header" style="color: #ffc107;">🟡 Teljesített + Hiányos adatok</div>
            <div class="matrix-val" style="color: #ffc107;">{len(group2)} fő</div>
            <div class="matrix-sub">Kiküldhető lenne az érmük, de hiányzik a telefonszámuk vagy a csomagpontjuk.</div>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<div style='margin-top: 1.5rem;'></div>", unsafe_allow_html=True)
        
        st.markdown(f"""
        <div class="matrix-card" style="background: rgba(220, 53, 69, 0.02); border-color: rgba(220, 53, 69, 0.2);">
            <div class="matrix-header" style="color: #dc3545;">🔴 Nem teljesített + Nincs adata</div>
            <div class="matrix-val" style="color: #dc3545;">{len(group4)} fő</div>
            <div class="matrix-sub">Még nem teljesítették a kihívást és semmilyen adatot nem adtak meg.</div>
        </div>
        """, unsafe_allow_html=True)
        
    st.markdown("---")

    # Filter section
    st.header("🔎 Keresés és Részletes Csomaglista")
    
    search_query = st.text_input("Keresés név vagy e-mail alapján:", "").strip().lower()
    
    # Tab layout
    t_packages, t1, t2, t3, t4, t_all = st.tabs([
        "📦 Szállítandó Csomagok (Érem kalkuláció)",
        f"🟢 Kész a postázásra ({len(group1)})", 
        f"🟡 Hiányos teljesítők ({len(group2)})", 
        f"⚪ Nem teljesített + Van adat ({len(group3)})", 
        f"🔴 Nem teljesített + Nincs adat ({len(group4)})",
        "📋 Összes nevező"
    ])
    
    def display_runners_table(runners_list):
        if not runners_list:
            st.info("Nincsenek egyező nevezők ebben a csoportban.")
            return
            
        filtered = runners_list
        if search_query:
            filtered = [r for r in runners_list if search_query in r["name"].lower() or search_query in r["email"].lower()]
            
        if not filtered:
            st.warning("Nincs találat a keresési feltételekre.")
            return
            
        code_to_name = {code: name for name, code in lockers_list}
        
        data = []
        for r in filtered:
            status_badge = "🟢 Kész" if r["all_data_ok"] else f"❌ Hiányzó: {', '.join(r['missing_fields'])}"
            if not r["completed"]:
                status_badge = "⏳ Még nem teljesített"
                
            shipped_badge = "✅ Kiküldve" if r["shipped"] else "❌ Nincs feladva"
            
            official_locker_name = "Nincs"
            if r["locker_code"]:
                official_locker_name = code_to_name.get(r["locker_code"], "Ismeretlen kód")
            
            data.append({
                "Sor": r["row_index"],
                "Név": r["name"],
                "Email": r["email"],
                "Telefonszám": r["phone"] or "Nincs",
                "Megadott Csomagpont": r["point_description"] or "Nincs",
                "Feloldott Kód": r["locker_code"] or "Nincs",
                "Hivatalos Csomagpont Név": official_locker_name,
                "Teljesítés dátuma": r["completion_date"] or "Még nem",
                "Státusz": status_badge,
                "Feladva": shipped_badge,
                "Megjegyzés (Sheet)": r["comment"] or "-",
                "Megjegyzés (Tally)": r["tally_comment"] or "-"
            })
            
        df = pd.DataFrame(data)
        st.dataframe(df, use_container_width=True)

    with t_packages:
        st.subheader("📦 Tervezett Foxpost Csomagok (Érem kalkuláció)")
        st.info("Ez a nézet a szállítási csoportok (együtt rendelők / túratársak) szerint összesíti a csomagokat, megmutatva a teljesítési állapotukat és a küldendő érmek számát.")
        
        if package_rows:
            filtered_packages = package_rows
            if search_query:
                filtered_packages = [p for p in package_rows if search_query in p["Címzett"].lower() or search_query in p["Recipient Email"].lower()]
                
            if filtered_packages:
                df_pkg = pd.DataFrame(filtered_packages)
                st.dataframe(df_pkg.drop(columns=["Recipient Email"]), use_container_width=True)
            else:
                st.warning("Nincs találat a keresési feltételekre.")
        else:
            st.info("Nincsenek aktív csomagok (legalább egy teljesítővel).")

    with t1:
        st.subheader("🟢 Kész a postázásra (Teljesített + Minden adat megvan)")
        st.info("Ezen ügyfelek adatai szerepelnek a generált tömeges Foxpost import Excel fájlban.")
        display_runners_table(group1)
        
    with t2:
        st.subheader("🟡 Hiányos teljesítők (Teljesített + Hiányos szállítási adatok)")
        st.warning("Velük fel kell venni a kapcsolatot e-mailben, vagy ki kell küldeni a szállítási adatok kérőjét (szállítási ping)!")
        display_runners_table(group2)
        
    with t3:
        st.subheader("⚪ Nem teljesített, de megadta az adatait")
        st.info("Ezek a futók már regisztrálták a szállítási adataikat előre, de a futást még nem igazolták be.")
        display_runners_table(group3)
        
    with t4:
        st.subheader("🔴 Nem teljesített és nincs semmilyen adata")
        display_runners_table(group4)
        
    with t_all:
        st.subheader("📋 Összes nevező listája")
        display_runners_table(all_runners)

    st.markdown("---")

    # Action panel
    st.header("⚙️ Logisztikai Műveletek")
    
    col_act1, col_act2 = st.columns(2)
    
    with col_act1:
        st.subheader("Foxpost Excel Generálás")
        st.write("Generáld le a tömeges Foxpost import Excel fájlt a legfrissebb Google Sheets adatok alapján.")
        
        if st.button("🚀 Foxpost Excel Generálása", use_container_width=True):
            with st.spinner("Generálás folyamatban..."):
                try:
                    # Filter for ready runners (completed, not shipped, and having all data correct)
                    ready_runners = [r for r in all_runners if r["completed"] and not r["shipped"] and r["all_data_ok"]]
                            
                    # Grouping ONLY based on the "együtt küldve" (primary_buyer) column relationship
                    shipments = {}
                    for r in ready_runners:
                        key = r["primary_buyer"] if (r["is_sub_order"] and r["primary_buyer"]) else r["email"]
                        if key not in shipments:
                            primary_runner = next((x for x in raw_runners if x["email"].lower() == key.lower()), None)
                            if primary_runner:
                                recip_name = primary_runner["name"]
                                recip_email = primary_runner["email"]
                                recip_phone = primary_runner["phone"]
                                point_desc = primary_runner["point_description"]
                                locker_code = primary_runner["locker_code"]
                            else:
                                recip_name = r["name"]
                                recip_email = r["email"]
                                recip_phone = r["phone"]
                                point_desc = r["point_description"]
                                locker_code = r["locker_code"]
                                
                            shipments[key] = {
                                "recipient_name": recip_name,
                                "recipient_email": recip_email,
                                "recipient_phone": recip_phone,
                                "point_description": point_desc,
                                "locker_code": locker_code,
                                "runners": [],
                                "rows": []
                            }
                        shipments[key]["runners"].append({"name": r["name"], "email": r["email"]})
                        shipments[key]["rows"].append(r["row_index"])

                    # Filter shipments to ensure they have all required data
                    valid_shipments = {}
                    for key, s in shipments.items():
                        if s["recipient_phone"] and s["recipient_email"] and s["locker_code"]:
                            valid_shipments[key] = s
                        
                    # Write workbook
                    template_wb = openpyxl.load_workbook(XLSX_TEMPLATE_PATH)
                    example_sheet = template_wb['Példa adatok']
                    
                    max_r = example_sheet.max_row
                    if max_r >= 2:
                        example_sheet.delete_rows(2, max_r - 1)
                        
                    row_count = 2
                    for s in valid_shipments.values():
                        # Dedup by name and email
                        unique_runners = []
                        seen_runners = set()
                        for runner in s["runners"]:
                            r_key = (runner["name"].lower().strip(), runner["email"].lower().strip())
                            if r_key not in seen_runners:
                                seen_runners.add(r_key)
                                unique_runners.append(runner)
                                
                        # Only display runners who have a different email than the recipient
                        display_runners = [x["name"] for x in unique_runners if x["email"].lower().strip() != s["recipient_email"].lower().strip()]
                        
                        if display_runners:
                            recip_name = f"{s['recipient_name']} (+ {', '.join(display_runners)})"
                        else:
                            recip_name = s['recipient_name']
                            
                        ph_val = s['recipient_phone']
                        if ph_val:
                            ph_val = int(ph_val)
                            
                        example_sheet.cell(row=row_count, column=1, value=recip_name)
                        example_sheet.cell(row=row_count, column=2, value=ph_val)
                        example_sheet.cell(row=row_count, column=3, value=s['recipient_email'])
                        example_sheet.cell(row=row_count, column=4, value=s['locker_code'])
                        example_sheet.cell(row=row_count, column=8, value=0)
                        example_sheet.cell(row=row_count, column=9, value='XS')
                        # Column 11 (Saját adatok) is left empty
                        
                        row_count += 1
                        
                    template_wb.save(OUTPUT_XLSX_PATH)
                    st.success(f"Sikeres generálás! Mentve ide: {OUTPUT_XLSX_PATH}")
                    
                except Exception as ex:
                    st.error(f"Sikertelen generálás: {ex}")
                    
    with col_act2:
        st.subheader("Csomagpont Párosító Ellenőrzés")
        st.write("Ezzel a modullal ellenőrizheted, hogy a Google Sheet-ben kézzel vagy űrlappal beírt szállítási címeket az algoritmus helyesen párosítja-e a Foxpost Locker listájával.")
        
        st.markdown(f"Összesen **{len(lockers_list)}** regisztrált Foxpost automata érhető el a rendszerben.")
        
        # Test custom match
        test_txt = st.text_input("Tesztelj le egy tetszőleges címszöveget:", "Foxpost Z-box Mr. Mosi Regős u. 9")
        if test_txt:
            res_code = resolve_locker_code(test_txt, lockers_list)
            if res_code:
                # Find name of this code
                off_name = next((name for name, code in lockers_list if code == res_code), "Ismeretlen név")
                st.success(f"Találat! Kód: **{res_code}** ({off_name})")
            else:
                st.error("Nincs találat erre a címszövegre.")

if __name__ == "__main__":
    main()
